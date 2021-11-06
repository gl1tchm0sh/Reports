import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import datetime
import time
from Merge_results import merge as full_data

def export_to_excel(filepath):

    """
    Exportar datos a excel
    Argumentos:
    filepath - Ubicacion y nombre donde se quiere guardar el archivo
    """
    # Las dos variables siguientes definen los dos formatos utlilizados mas adelante para
    # establecer los bordes de las celdas
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    thick_left = Border(left=Side(border_style='thick', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))

    data = full_data()
    # Elimina 'Servicio tecnico' del listado en caso de que apareciera
    if 'SERVICIO TECNICO' in data:
        del data['SERVICIO TECNICO']

    vdr_zona = {
        "JORGE ESTEBAN AYERDI":1, "CARLOS FEDERICO MAUMUS":3, "HERNAN JAVIER CORBALAN":4,
        "LUIS MARIANO PEREYRA":5, "JUAN JOSE STAIN":6, "SEBASTIAN IGNACIO COUTO":7,
        "JOSE LUIS MAURO":8, "LUIS EDGARDO CASTRO":9, "FERNANDO MARIA RODRIGUEZ VILLAMIL":10,
        "PABLO LUCIANO JAICOVSKY":11, "SEBASTIAN ERNESTO SIRI":13, "MARCELO LEONARDO ZANFAGNINI":18,
        "ROBERTO ALEJANDRO BAEZ":19, "HERNAN PARRUCCI ARIEL":21, "WALTER HORACIO PICCARDI":22,
        "RICARDO MANUEL BETORZ":23, "JUAN PABLO DILUCA":24, "PABLO SEBASTIAN RODRIGUEZ":25,
        "GONZALO ROBERTO SAUL":26, "SILVIO MARCEL COSTA":27, "MATIAS CARLOS BETTINI":33,
        "GUILLERMO MAURO COMBA":35, "EMILIANO IGNACIO SANZ":38, "JORGE ANTONIO DI PAULI":40,
        "MAURICIO SEBASTIAN RODRIGUEZ":42, "SERGIO ADRIAN DOMINGO VERSACE":45,
        "CRISTIAN ARIEL RODRIGUEZ":49, "ARIEL DIAZ":81
    }
    # Carga Template.xlsx para usarlo de formato base
    wb = openpyxl.load_workbook('Template.xlsx')
    template = wb['Hoja1']
    sheet = wb.copy_worksheet(template)

    Date = datetime.date.today()
    Month = Date.strftime('%B')

    sheet.title = f'{Month}'
    template.sheet_state = 'hidden'
    wb.active = sheet


    # Toma los objetivos desde el Template. Si se modifican los valores en este, se
    # establecera como objetivo el valor que se guarde.
    o_gedore = int(sheet.cell(row=2, column=8).value)
    o_atlas = int(sheet.cell(row=2,column=10).value)
    o_tyrolit = int(sheet.cell(row=1,column=12).value)
    o_bosch = int(sheet.cell(row=1,column=14).value)
    o_bovenau = int(sheet.cell(row=2,column=16).value)

    # Este es el loop principal, donde se da formato y se empujan los datos a cada celda
    # Se hardcodearon las columnas para agilizar el lanzamiento, pero se puede generar otra función que tome
    # los objetivos del template, y lo agregue a data, permitiendo un loop mas ordenado
    for rowno, vdr in enumerate(data, start = 5):
        sheet.cell(row=rowno, column=1).value = vdr_zona[vdr]
        sheet.cell(row=rowno, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=rowno, column=1).font = Font(b=True,i=True)
        sheet.cell(row=rowno, column=2).value = str(vdr).lower().title()
        sheet.cell(row=rowno, column=2).font = Font(b=True, i=True)

        #Info de columnas 3 y 4 sale de Persat
        #Las columnas de los porcentajes son columnas calculadas luego de empujar el valor de cada celda a continuación
        if 'Desarrollo' in data[vdr]:
            sheet.cell(row=rowno, column=5).value = round(float(data[vdr]['Desarrollo']),2)
            sheet.cell(row=rowno, column=7).value = str(int((float(data[vdr]['Desarrollo']) * 100) // float(data[vdr]['Total']))) + '%'
        sheet.cell(row=rowno, column=6).value = data[vdr]['Total']
        sheet.cell(row=rowno, column=18).value = data[vdr]['TotalA']
        if 'Gedore' in data[vdr]:
            sheet.cell(row=rowno, column=8).value = round(float(data[vdr]['Gedore']),2)
            sheet.cell(row=rowno, column=9).value = str(int((float(data[vdr]['Gedore']) * 100) // o_gedore)) + '%'
        if 'Atlas' in data[vdr]:
            sheet.cell(row=rowno, column=10).value = round(float(data[vdr]['Atlas']),2)
            sheet.cell(row=rowno, column=11).value = str(int((float(data[vdr]['Atlas']) * 100) // o_atlas)) + '%'
        if 'Tyrolit' in data[vdr]:
            sheet.cell(row=rowno, column=12).value = round(float(data[vdr]['Tyrolit']),2)
            sheet.cell(row=rowno, column=13).value = str(int((float(data[vdr]['Tyrolit']) * 100) // o_tyrolit)) + '%'
        if 'Bosch' in data[vdr]:
            sheet.cell(row=rowno, column=14).value = round(float(data[vdr]['Bosch']),2)
            sheet.cell(row=rowno, column=15).value = str(int((float(data[vdr]['Bosch']) * 100) // o_bosch)) + '%'
        if 'Bovenau' in data[vdr]:
            sheet.cell(row=rowno, column=16).value = round(float(data[vdr]['Bovenau']),2)
            sheet.cell(row=rowno, column=17).value = str(int((float(data[vdr]['Bovenau']) * 100) // o_bovenau)) + '%'

        col_list = [7,9,11,12,13,15,17]
        for colno in col_list:
            sheet.cell(row=rowno, column=colno).alignment = Alignment(horizontal='center')


    #Ajusta el tamaño de cada columna según el máximo ancho de su contenido.
    # Esta funcion es valida para cualquier otro script de openpyxl o de excel en general
    # se puede separar en una nueva funcion.
    for colno in range(1, sheet.max_column +1):
        column_max_width = 0
        for rowno in range (1, sheet.max_row +1):
            current_cell_width = len(str(sheet.cell(row = rowno, column = colno).value))
            if current_cell_width > column_max_width:
                column_max_width = current_cell_width+2
        if column_max_width < 10:
            sheet.column_dimensions[get_column_letter(colno)].width = 10
        else:
            sheet.column_dimensions[get_column_letter(colno)].width = column_max_width

    #Inserta la fecha como cabecera de la tabla, y da formato a todas las cabeceras junto con esta.

    formatted_month = ''
    for char in Month:
        formatted_month += char + ' '

    sheet.cell(row = 1, column = 1).value = formatted_month

    sheet.cell(row=1, column=1).font = Font(bold=False, color="000000")
    sheet.cell(row=1, column=1).fill =  PatternFill('solid',fgColor='E2EFDA')
    sheet.cell(row=1, column=1).font += Font(size=18)
    sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal = 'center')
    for row in range(5,sheet.max_row+1):
        col_list = [2,9,11,13,15,17]
        for colno in col_list:
            value = str(sheet.cell(row=row, column=colno).value).replace('%','')
            try:
                value = int(value)
            except:
                continue
            value = int(value)
            if value >= 100:
                sheet.cell(row=row, column=colno).fill = PatternFill(start_color='FFF2CC', end_color='A9D08E', fill_type='gray125')
        value = sheet.cell(row=row, column=18).value
        if value > 0:
            sheet.cell(row=row, column=18).fill = PatternFill(start_color='FFF2CC', end_color='A9D08E', fill_type='gray125')
        sheet.cell(row=row, column=17).border = thin_border
        sheet.cell(row=row, column=1).border = thin_border
        sheet.cell(row=row, column=2).border = thin_border
        sheet.cell(row=row, column=5).border = thick_left
        sheet.cell(row=row, column=3).border = thick_left

        for colno in range(6,sheet.max_column-1):
            sheet.cell(row=row, column=colno).border = Border(left=Side(border_style='thin', color='FF000000'))
            if row == sheet.max_row:
                sheet.cell(row=row, column=colno).border = Border(bottom=Side(border_style='thin', color='FF000000'),
                                                                  left=Side(border_style='thin', color='FF000000'))

    wb.save(filepath)
    print("xlsx created succesfully")



start_time = time.time()


#Modificar el filepath en caso de ser necesario, siempre con doble barra \\, para evitar errores de sintaxis
guardar_en = "C:\\Ejemplo.xlsx"


export_to_excel(guardar_en)
print("--- %s seconds ---" % round((time.time() - start_time),2))
