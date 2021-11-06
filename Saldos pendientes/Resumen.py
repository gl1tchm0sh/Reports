import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def summarize(wb, data):
    cabeceras = ['Cliente', 'Total original', 'Total adeudado', 'Total vencido']
    client_list = list(data.keys())
    if 'Resumen' not in wb.sheetnames:
        wb.create_sheet('Resumen')
    sheet = wb['Resumen']
    wb.active = sheet
    sheet.delete_cols(1, sheet.max_column)

    for colno, heading in enumerate(cabeceras, start=1):
        sheet.cell(row=1, column=colno).value = heading

    for rowno, client in enumerate(client_list, start=2):
        sheet.cell(row=rowno, column=1).value = client
        sheet.cell(row=rowno, column=2).value = data[client]['total']
        sheet.cell(row=rowno, column=3).value = data[client]['deuda']
        sheet.cell(row=rowno, column=4).value = data[client]['vencido']

    for colno in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=colno).font = Font(bold=True, color="FFFFFF")
        sheet.cell(row=1, column=colno).fill = PatternFill('solid', fgColor='0000FF')

    for colno in range(1, sheet.max_column + 1):
        column_max_width = 0
        for rowno in range(1, sheet.max_row + 1):
            current_cell_width = len(str(sheet.cell(row=rowno, column=colno).value))
            if current_cell_width > column_max_width:
                column_max_width = current_cell_width
        sheet.column_dimensions[get_column_letter(colno)].width = column_max_width

    sticky = sheet['B2']
    sheet.freeze_panes = sticky



