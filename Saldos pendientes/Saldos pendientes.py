import openpyxl
import pgconnection
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import datetime
from Resumen import summarize

def export_to_excel(connection, query_string, headings, filepath):
    """
        Exportar datos a excel, generando una hoja por cada vendedor
        que tenga clientes con movimientos registrado.
        Argumentos:
        connection - Una conexión abierta de psycopg2
        query_string - La query que se envía a la BD
        headings - Cabeceras para el excel resultante
        filepath - Ubicacion y nombre donde se quiere guardar el archivo
        """
    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    data.sort()
    # Se opta por convertir la tupla en lista, para poder
    # eliminar los datos a medida que son empujados al excel
    # reduciendo casi a la mitad los tiempos de procesamiento.
    data = list(data)
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.active
    vdrs_list = ['']
    Date = datetime.date.today()

    y_now = datetime.date.today().year
    m_now = datetime.date.today().month
    d_now = datetime.date.today().day

    # Genera la lista de vendedores a partir de los datos
    for rowno, row in enumerate(data):
        if row[0] not in vdrs_list:
            vdrs_list.append(row[0])

    for vdr in vdrs_list:
        summary = {}
        sheet.delete_cols(1,sheet.max_column) # Limpia la hoja

        # El indice de la hoja es 1, por lo que start se asigna como 1
        # para empujar las cabeceras a la hoja
        for colno, heading in enumerate(headings, start = 1):
            sheet.cell(row = 1, column = colno).value = heading

        # En este caso start = 2 para saltear la primer linea (cabeceras)
        for rowno, row in enumerate(data, start = 2):
            last_rowno = 0
            if row[0] != vdr:
                last_rowno = rowno-2
                del(data[0:last_rowno])
                break
            cliente = ''
            adeuda = 0
            total_impago = 0
            total_vencido = 0
            for colno, cell_value in enumerate(row, start = 1):
                if colno == 2:
                    cliente = cell_value
                elif colno == 4:
                    adeuda = float(cell_value.replace('$','').replace('.','').replace(',','.').strip())
                elif colno == 5:
                    total_impago = float(cell_value.replace('$','').replace('.','').replace(',','.').strip())
                elif colno == 6 or colno == 7:
                    year = int(cell_value[0:4])
                    month = int(cell_value[4:6])
                    day = int(cell_value[6:8])
                    cell_value = f'{day}/{month}/{year}'
                    if colno == 6:
                        if year > y_now or year == y_now and month < m_now or year == y_now and month == m_now and day < d_now:
                            importe = sheet.cell(row=rowno, column=4).value
                            total_vencido = float(importe.replace('$','').replace('.','').replace(',','.').strip())

                sheet.cell(row = rowno, column = colno).value = cell_value
            if cliente not in summary:
                summary[cliente] = {}
                summary[cliente]['deuda'] = adeuda
                summary[cliente]['total'] = total_impago
                summary[cliente]['vencido'] = total_vencido
            else:
                summary[cliente]['deuda'] += adeuda
                summary[cliente]['deuda'] = round(summary[cliente]['deuda'], 2)
                summary[cliente]['total'] += total_impago
                summary[cliente]['total'] = round(summary[cliente]['total'], 2)
                summary[cliente]['vencido'] += total_vencido
                summary[cliente]['vencido'] = round(summary[cliente]['vencido'],2)




        #Inserta la fecha como cabecera de la tabla, y da formato a todas las cabeceras junto con esta.
        sheet.delete_cols(1)

        ## Ajusta el tamaño de cada columna según el máximo ancho de su contenido.
        for colno in range(1, sheet.max_column + 1):
            column_max_width = 0
            for rowno in range(1, sheet.max_row + 1):
                current_cell_width = len(str(sheet.cell(row=rowno, column=colno).value))
                if current_cell_width > column_max_width:
                    column_max_width = current_cell_width
            sheet.column_dimensions[get_column_letter(colno)].width = column_max_width

        sticky = sheet['B3']
        sheet.freeze_panes = sticky

        sheet.insert_rows(1)
        sheet.merge_cells(f"A1:{get_column_letter(sheet.max_column)}1")
        sheet.cell(row = 1, column = 1).value = str(vdr) + ' '*50 +'Listado general de saldos al ' + str(Date)

        sheet.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        sheet.cell(row=1, column=1).fill =  PatternFill('solid',fgColor='0000FF')
        sheet.cell(row=1, column=1).font += Font(size=20)
        sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal = 'left')
        for colno in range(1,sheet.max_column+1):
            sheet.cell(row=2, column=colno).font = Font(bold=True, color ="FFFFFF")
            sheet.cell(row=2, column=colno).fill = PatternFill('solid', fgColor='0000FF')

        summarize(wb,summary)
        wb.active = sheet
        filepath_new = filepath + str(vdr) + '_1.xlsx'
        wb.save(filepath_new)
        print(f"{vdr}_1.xlsx created succesfully")


query = """SELECT   (CASE
          WHEN ALIAS_0.NOMBREOPERADORASOCIADO = 'ADELQUIS NAHUEL TOMASSINI'
          THEN 'HERNAN PARRUCCI ARIEL'
          WHEN ALIAS_0.NOMBREOPERADORASOCIADO = 'MAXIMILIANO ALDERETE'
          THEN 'GUILLERMO MAURO COMBA'
          WHEN ALIAS_0.NOMBREOPERADORASOCIADO = 'PABLO NORBERTO MATALONI'
          THEN ''
          WHEN ALIAS_0.NOMBREOPERADORASOCIADO = 'JOSE MARIO KUNTZ'
          THEN 'HERNAN JAVIER CORBALAN'
          WHEN ALIAS_0.NOMBREOPERADORASOCIADO = 'RUBEN JORGE BRAVI'
          THEN 'MARCELO LEONARDO ZANFAGNINI'
          ELSE
          ALIAS_0.NOMBREOPERADORASOCIADO END) ALIAS_0_NOMBREOPERADORASOCIADO, 
        ALIAS_1.NOMBRE ALIAS_1_NOMBRE, ALIAS_0.DESCRIPCION ALIAS_0_DESCRIPCION,
        ALIAS_2.NOMBRE ALIAS_2_NOMBREtotalpendiente, ALIAS_3.NOMBRE ALIAS_3_NOMBREtotaloriginal, ALIAS_0.FECHAESTIMADA ALIAS_0_FECHAESTIMADA, 
        ALIAS_0.FECHAEMISION ALIAS_0_FECHAEMISION, ALIAS_5.NOMBRE ALIAS_5_NOMBRE, 
        ALIAS_0.NOMCLASIFICADOR ALIAS_0_NOMCLASIFICADOR
        
        FROM   V_COMPROMISOPAGO ALIAS_0  
        
        LEFT OUTER JOIN V_OPERADORCOMERCIAL ALIAS_15 ON ALIAS_0.OPERADORCOMERCIAL_ID = ALIAS_15.ID   
        LEFT OUTER JOIN V_PERSONA ALIAS_1 ON ALIAS_15.ENTEASOCIADO_ID = ALIAS_1.ID   
        LEFT OUTER JOIN V_VALOR ALIAS_2 ON ALIAS_0.SALDO_ID = ALIAS_2.ID   
        LEFT OUTER JOIN V_VALOR ALIAS_3 ON ALIAS_0.IMPORTETOTAL_ID = ALIAS_3.ID   
        LEFT OUTER JOIN V_ESTADO ALIAS_4 ON ALIAS_0.ESTADO_ID = ALIAS_4.ID   
        LEFT OUTER JOIN V_TIPOPAGO ALIAS_5 ON ALIAS_0.TIPOPAGO_ID = ALIAS_5.ID   
        LEFT OUTER JOIN V_UNIDADFINANCIERA ALIAS_6 ON ALIAS_0.SALDO2_UNIDADVALORIZACION_ID = ALIAS_6.ID   
        LEFT OUTER JOIN V_UNIDADFINANCIERA ALIAS_7 ON ALIAS_0.IMTOTAL2_UNIDADVALORIZACION_ID = ALIAS_7.ID   
        LEFT OUTER JOIN V_IMPUTACIONCONTABLE ALIAS_8 ON ALIAS_0.TRIMPUTACIONCONTABLE_ID = ALIAS_8.ID   
        LEFT OUTER JOIN V_CENTROCOSTOS ALIAS_9 ON ALIAS_0.CENTROCOSTOS_ID = ALIAS_9.ID   
        LEFT OUTER JOIN V_TIPOTRANSACCION ALIAS_10 ON ALIAS_0.TIPOTRANSACCION_ID = ALIAS_10.ID   
        LEFT OUTER JOIN V_UNIDADNEGOCIO ALIAS_11 ON ALIAS_0.UNIDADNEGOCIO_ID = ALIAS_11.ID   
        LEFT OUTER JOIN V_SUBSUCURSAL ALIAS_12 ON ALIAS_0.SUBSUCURSAL_ID = ALIAS_12.ID   
        LEFT OUTER JOIN V_AREARESPONSABILIDAD ALIAS_13 ON ALIAS_0.AREARESP_ID = ALIAS_13.ID   
        LEFT OUTER JOIN V_PROYECTO ALIAS_14 ON ALIAS_0.PROYECTO_ID = ALIAS_14.ID  
        
 WHERE  ALIAS_0.BO_PLACE_ID IS NOT NULL   AND  ALIAS_0.SALDADO = 'F'  
        AND  ALIAS_0.NIVEL = 1  
        AND  ALIAS_0.CPCOMPRAS = 'F'  
        AND ALIAS_0.NOMBREOPERADORASOCIADO NOT LIKE ''
        AND (ALIAS_10.DESCRIPCION LIKE 'Factura de Venta' OR ALIAS_10.DESCRIPCION LIKE 'Saldos Clientes - FV y ND'  OR ALIAS_10.DESCRIPCION LIKE 'Nota de Débito Venta x Cheque Rechazado' OR ALIAS_10.DESCRIPCION LIKE '%Débito%')
        AND   ( ALIAS_0.CLASIFICADOR_ID IN ( '{EAD5B984-766F-44B3-AB16-805AA13DE6C7}' , '{767608F0-952F-4F04-8C03-6F29B8F1299A}' )   OR   ALIAS_0.CLASIFICADOR_ID IS NULL  )
        ORDER BY ALIAS_0.NOMBREOPERADORASOCIADO, ALIAS_1.NOMBRE, ALIAS_0.FECHAESTIMADA
 """

cabeceras = ['Vendedor', 'Cliente', 'Comprobante', 'Saldo Pendiente', 'Saldo Original', 'Fecha de vencimiento', 'Fecha de emisión', 'Tipo de Pago', 'Esquema Operativo']


#Modificar el filepath en caso de ser necesario, siempre con doble barra \\, para evitar errores de sintaxis
guardar_en = "C:\\Saldos.xlsx"


#Revisar el script pgconnection en caso de necesitar modificar datos de conexion
conn = pgconnection.get_connection("BD")

export_to_excel(conn, query, cabeceras, guardar_en)
