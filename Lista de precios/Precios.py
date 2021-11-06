import openpyxl
import pgconnection
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import datetime
from Codigosalternativos import Get_query_list as c_alt

def export_to_excel(connection, query_string, headings, filepath):

    """
    Exportar querys a excel usando psycopg2 (En el script pgconnection).
    Argumentos:
    connection - un psycopg2 abierto
    query_string - SQL para tomar los datos
    headings - Lista de strings para tomar las cabeceras
    filepath - Ubicacion y nombre donde se quiere guardar el archivo
    """

    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.active

    # Los indices de fila y columna en la hoja de calculo comienzan en 1
    # Asi que se utiliza "start = 1" en la funcion enumerate
    # Para no tener que agregarlo manualmente en la funcion
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    # Query de para tomar los codigos alternativos
    query_alt = "SELECT ALIAS_0.CODIGO ALIAS_0_CODIGO, ALIAS_1.CODIGOALTERNATIVO FROM V_PRODUCTO ALIAS_0 LEFT OUTER JOIN V_CODIGOALTERNATIVO ALIAS_1 ON ALIAS_1.BO_PLACE_ID = ALIAS_0.CODIGOSALTERNATIVOS_ID LEFT OUTER JOIN V_PROVEEDOR ALIAS_2 ON ALIAS_2.ID = ALIAS_1.OPERADORCOMERCIAL_ID WHERE ALIAS_2.BO_PLACE_ID = '{77B039B8-4A31-11D5-B607-0050DAC017BE}' AND  ALIAS_2.ACTIVESTATUS = 0 AND  ALIAS_2.TIPOOBJETOESTATICO_ID = '{066AE32F-8880-4B2C-AC0C-5D97C57FCFF6}' ORDER BY ALIAS_0.CODIGO"
    alternativos = c_alt(connection, query_alt)

    # En este caso "start = 2" es para saltear la cabecera, comenzamos en la segunda linea
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

            #Si la columna es la 'G', le asigna el formato apropiado de numero
            if sheet.cell(row = rowno, column = colno).value == sheet[f'G{rowno}'].value and sheet.cell(row = rowno, column = colno).value != None:
                modify = str(sheet.cell(row=rowno, column=colno).value)
                modify = modify.replace('$', '').replace(',', '.').replace(' ', '')
                sheet.cell(row=rowno, column=colno).number_format = '[$$-es-AR] #.##;-[$$-es-AR] #.##' #Formato coincide exactamente con el de excel, se puede sacar el codigo de ahí
                sheet.cell(row = rowno, column = colno).value = round(float(modify),2)
                val_60d = round(float(modify) * 1.0378,2) # Calcula el precio resultante si lo abona a 60 dias
                sheet.cell(row=rowno, column=12).value=val_60d
                val_epa = round(float(modify) * 0.95,2) # Calcula el precio resultante si lo abona en EPA
                sheet.cell(row=rowno, column=13).value = val_epa

    #busca el código en el listado "alternativos" y si existe, pega el codigo alternativo, columna K (11) hardcodeada
    for rowno in range(2,sheet.max_row):
        code = sheet.cell(row=rowno, column=2).value
        if code in alternativos:
            sheet.cell(row=rowno, column=11).value = alternativos[code]
            del alternativos[code]


    #Ajusta el tamaño de cada columna según el máximo ancho de su contenido.
    for colno in range(1, sheet.max_column +1):
        column_max_width = 0
        for rowno in range (1, sheet.max_row +1):
            current_cell_width = len(str(sheet.cell(row = rowno, column = colno).value))
            if current_cell_width > column_max_width:
                column_max_width = current_cell_width
        sheet.column_dimensions[get_column_letter(colno)].width = column_max_width

    #Inserta la fecha como cabecera de la tabla, y da formato a todas las cabeceras junto con esta.
    Date = datetime.date.today()
    sheet.insert_rows(1)
    sheet.merge_cells(f"A1:{get_column_letter(sheet.max_column)}1")
    sheet.cell(row = 1, column = 1).value = Date

    sheet.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=1, column=1).fill =  PatternFill('solid',fgColor='0000FF')
    sheet.cell(row=1, column=1).font += Font(size=20)
    sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal = 'center')
    for colno in range(1,sheet.max_column+1):
        sheet.cell(row=2, column=colno).font = Font(bold=True, color ="FFFFFF")
        sheet.cell(row=2, column=colno).fill = PatternFill('solid', fgColor='0000FF')

    wb.save(filepath)
    print("xlsx created succesfully")




query = """SELECT ALIAS_6.NOMBRERUBRO ALIAS_6_NOMBRERUBRO, ALIAS_0.CODIGO ALIAS_0_CODIGO, ALIAS_0.DESCRIPCION ALIAS_0_DESCRIPCION, ALIAS_0.INICIAL2_IMPORTE ALIAS_0_INICIAL2_IMPORTE, ALIAS_2.NOMBRE ALIAS_2_NOMBRE, ALIAS_0.PORCENTAJE ALIAS_0_PORCENTAJE, ALIAS_0.VALOR2_NOMBRE ALIAS_0_VALOR2_NOMBRE, ALIAS_3.NOMBRE ALIAS_3_NOMBRE,itemposicionadorimpuestos.coeficiente,  ALIAS_5.UNIDMINIMCOMER ALIAS_5_UNIDMINIMCOMER
FROM V_PRECIO ALIAS_0
LEFT OUTER JOIN V_UNIDADFINANCIERA ALIAS_2 ON ALIAS_0.VALOR2_UNIDADVALORIZACION_ID = ALIAS_2.ID
LEFT OUTER JOIN V_UNIDADMEDIDA ALIAS_3 ON ALIAS_0.DCANTIDAD2_UNIDADMEDIDA_ID = ALIAS_3.ID
LEFT OUTER JOIN V_UNIDADMEDIDA ALIAS_4 ON ALIAS_0.HCANTIDAD2_UNIDADMEDIDA_ID = ALIAS_4.ID 
LEFT OUTER JOIN v_PRODUCTO ALIAS_5 ON ALIAS_5.ID = REFERENCIA_ID 
LEFT OUTER JOIN V_RUBRO ALIAS_6 ON ALIAS_5.RUBRO_ID = ALIAS_6.ID
        LEFT JOIN posicionadorimpuestos ON posicionadorimpuestos.ID = alias_5.POSICIONADORIMPUESTOS_ID
        LEFT JOIN itemposicionadorimpuestos ON itemposicionadorimpuestos.BO_PLACE_ID = posicionadorimpuestos.itemsposicionadorimpuestos_id
        LEFT JOIN definicionimpuesto ON definicionimpuesto.ID = itemposicionadorimpuestos.DEFINICIONIMPUESTO_id -- AND definicionimpuesto.IMPUESTO_ID ='a0b20948-c746-4b65-b42e-91df0addbd52'
        LEFT JOIN IMPUESTO ON IMPUESTO.ID = definicionimpuesto.IMPUESTO_ID
WHERE IMPUESTO.NOMBRE = 'IVA TASA'
        AND ALIAS_0.BO_PLACE_ID = '{15430301-BD4D-4501-8393-A3E02F149542}'   
        AND  ALIAS_0.ACTIVESTATUS <> 2 
        AND  ALIAS_0.CODIGO NOT LIKE 'MO.%'
ORDER BY ALIAS_6_NOMBRERUBRO, ALIAS_0_CODIGO  LIMIT 40000"""

cabeceras = ["Rubro", "Codigo", "Descripcion", "Importe inicial", "Moneda", "Descuento %", "Importe final", "U. Precio", "IVA", "Bulto Minorista", "Codigo Alternativo", "60 Dias", "EPA"]


#Modificar el filepath en caso de ser necesario, siempre con doble barra \\, para evitar errores de sintaxis
guardar_en = "C:\\Carpeta\\Archivo.xlsx"


#Revisar el script pgconnection en caso de necesitar modificar datos de conexion
conn = pgconnection.get_connection("DB")

export_to_excel(conn, query, cabeceras, guardar_en)
