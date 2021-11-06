import openpyxl
from articulosetman import main_process as articulos
from Retrieve_rbrs import retrieve
from convert_to_csv import conversion

def export_to_excel_1(filepath_write, filepath_read):

    """
    Exportar un listado de parametros obtenidos segun el listado
    provisto por un excel
    Argumentos:
    filepath_write - Ubicacion y nombre donde se quiere guardar el archivo
    filepath_read - Ubicacion y nombre donde se quieren tomar los datos
    """

    wb = openpyxl.Workbook()
    sheet = wb.active
    data = articulos(filepath_read)

    for rowno, row in enumerate(data, start=1):
        for colno, content in enumerate(row, start = 1):
            if colno == 4:
                content = int(content)
            elif colno == 5:
                content = round(float(content.replace("$", "").replace(",",".").strip()),2)
            elif colno == 6:
                content = content[0:8]
            sheet.cell(row =rowno, column = colno).value = str(content)
    sheet.delete_cols(1,1)
    sheet.insert_cols(4,1)
    for rowno, row in enumerate(data, start=1):
        sheet.cell(row=rowno, column=4).value = 1

    wb.save(filepath_write)
    print("xlsx created succesfully")

def export_to_excel_2(filepath_write, filepath_read):

    wb = openpyxl.Workbook()
    sheet = wb.active
    data = articulos(filepath_read)
    descuentos = retrieve(filepath_read)

    for rowno, row in enumerate(data, start=1):
        for colno, content in enumerate(row, start = 1):
            if colno == 4:
                content = int(content)
            elif colno == 5:
                content = round(float(content.replace("$", "").replace(",",".").strip()),2)
                rubro = int(sheet.cell(row=rowno, column=1).value)
                if rubro in descuentos:
                    content = content - (content * descuentos[rubro])
            elif colno == 6:
                content = content[0:8]
            sheet.cell(row =rowno, column = colno).value = str(content)

    sheet.delete_cols(1,1)
    sheet.insert_cols(4,1)

    for rowno, row in enumerate(data, start=1):
        sheet.cell(row=rowno, column=4).value = 1

    wb.save(filepath_write)
    print("xlsx created succesfully")



#Modificar el filepath en caso de ser necesario, siempre con doble barra \\, para evitar errores de sintaxis
leer_desde = 'RUBROS_DESCUENTOS.xlsx'
guardar_en = "C:\\articulos.xlsx"
guardar_en_2 = "C:\\costofinal.xlsx"

export_to_excel_1(guardar_en, leer_desde)
export_to_excel_2(guardar_en_2, leer_desde)

conversion(guardar_en,'articulos')
conversion(guardar_en_2,'costofinal')
