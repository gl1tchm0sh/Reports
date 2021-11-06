import openpyxl

"""
Utiliza openpyxl para abrir un archivo de excel que contenga
el listado de rubros y el respectivo descuento
Args:
filepath = Ubicacion del archivo de donde tomar los descuentos
"""
def retrieve(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    rbr_dsc = {}

    for rowno, row in enumerate(sheet, start=3):
        rbr = sheet.cell(row=rowno, column=1).value
        dsc = sheet.cell(row=rowno, column=7).value
        rbr_dsc[rbr] = dsc

    return rbr_dsc




