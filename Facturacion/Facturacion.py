import openpyxl
import pgconnection
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import datetime
import time


def export_to_excel(connection, query_string, headings, filepath):

    """
    Exportar querys a excel usando psycopg2 (En el script pgconnection).
    Argumentos:
    connection - un psycopg2 abierto
    query_string - SQL para tomar los datos
    headings - Lista de strings para tomar las cabeceras
    filepath - Ubicacion y nombre donde se quiere guardar el archivo
    """
    start_time = time.time()
    whole_time = time.time()

    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()

    print('Data fetched in:')
    print("--- %s seconds ---" % round((time.time() - start_time), 2))

    wb = openpyxl.Workbook()
    sheet = wb.active

    # Los indices de fila y columna en la hoja de calculo comienzan en 1
    # Asi que se utiliza "start = 1" en la funcion enumerate
    # Para no tener que agregarlo manualmente en la funcion
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    # En este caso "start = 2" es para saltear la cabecera, comenzamos en la segunda linea
    decimal_list = [20,21,22,24]
    comprobante_anterior = ''

    start_time = time.time()
    print('Pushing data')
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            if colno == 4:
                if cell_value == comprobante_anterior:
                    continue
            elif colno == 6:
                y = cell_value[0:4]
                m =cell_value[4:6]
                d = cell_value[6:8]
                cell_value = f'{d}/{m}/{y}'
            elif colno == 19 and len(str(cell_value)) > 0:
                cell_value = int(cell_value)
            elif colno in decimal_list and len(str(cell_value)) > 0:
                cell_value = round(float(cell_value),2)
            elif colno == 23 and len(str(cell_value)) > 0:
                cell_value = str(round(float(cell_value),2)) + '%'
            sheet.cell(row = rowno, column = colno).value = cell_value
        comprobante_anterior = row[3]
    print("Pushed data in : --- %s seconds ---" % round((time.time() - start_time), 2))

    start_time = time.time()
    print('Formatting data')

    #Ajusta el tamaño de cada columna según el máximo ancho de su contenido.
    for colno in range(1, sheet.max_column +1):
        column_max_width = 0
        for rowno in range (1, sheet.max_row +1):
            current_cell_width = len(str(sheet.cell(row = rowno, column = colno).value))
            if current_cell_width > column_max_width:
                column_max_width = current_cell_width
        sheet.column_dimensions[get_column_letter(colno)].width = column_max_width

    #Inserta la fecha como cabecera de la tabla, y da formato a todas las cabeceras junto con esta.
    Date = datetime.date.today()
    sheet.insert_rows(1)
    sheet.merge_cells(f"A1:{get_column_letter(sheet.max_column)}1")
    sheet.cell(row = 1, column = 1).value = Date

    sheet.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=1, column=1).fill =  PatternFill('solid',fgColor='0000FF')
    sheet.cell(row=1, column=1).font += Font(size=20)
    sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal = 'left')
    for colno in range(1,sheet.max_column+1):
        sheet.cell(row=2, column=colno).font = Font(bold=True, color ="FFFFFF")
        sheet.cell(row=2, column=colno).fill = PatternFill('solid', fgColor='0000FF')

    wb.save(filepath)
    print("Data formatted in --- %s seconds ---" % round((time.time() - start_time), 2))

    print("xlsx created succesfully in time:")
    print("--- %s seconds ---" % round((time.time() - whole_time), 2))





query = """
SELECT 
ALIAS_7.nombre UnidadOperativa,
ALIAS_3.codigo CodEsquemaOperativo,
ALIAS_3.descripcion EsquemaOperativo,
ALIAS_0.numerodocumento,
ALIAS_0.detalle,
ALIAS_0.fechaactual FechaEmision,
ALIAS_6.descripcion Flag,
ALIAS_0.estado,
ALIAS_0.nombreoriginante Originante,
ALIAS_0.nombredestinatario Destinatario,
ALIAS_24.descripcion Referencia,

ALIAS_30.nombre UnidadNegocio,
ALIAS_31.nombre TipoPago,
ALIAS_33.nombre2 ClasificadorCliente2,
ALIAS_33.nombre3 ClasificadorCliente3,
ALIAS_33.nombre5 ClasificadorCliente5,
ALIAS_35.nombre1 ClasificadorConceptoC1,
ALIAS_35.nombre2 ClasificadorConceptoC2,
ALIAS_24.cantidad2_cantidad Cantidad,
ALIAS_24.valor2_importe Valor,
ALIAS_24.total2_importe SubTotal,
ALIAS_24.preciounitariofinal PrecioUnitarioFinal,
ALIAS_24.porcentajebonificacion PorcentajeBonificacion,
ALIAS_24.suma2_importe ValorTotal,

ALIAS_32.cuit,
ALIAS_32.codigo CodCliente,

(SELECT P.NOMBRE POSICION
		   FROM CLIENTE C
		   LEFT JOIN V_POSICIONADORIMPUESTOS ON C.POSICIONADORIMPUESTOS_ID = V_POSICIONADORIMPUESTOS.ID
		   LEFT JOIN V_ITEMPOSICIONADORIMPUESTOS ON V_POSICIONADORIMPUESTOS.ITEMSPOSICIONADORIMPUESTOS_ID = V_ITEMPOSICIONADORIMPUESTOS.BO_PLACE_ID
		   LEFT JOIN V_POSICIONIMPUESTO P ON V_ITEMPOSICIONADORIMPUESTOS.POSICIONIMPUESTO_ID = P.ID
		   INNER JOIN V_DEFINICIONIMPUESTO ON V_ITEMPOSICIONADORIMPUESTOS.DEFINICIONIMPUESTO_ID = V_DEFINICIONIMPUESTO.ID
				AND V_DEFINICIONIMPUESTO.IMPUESTO_ID = '89C2361A-3F01-11D5-86AD-0080AD403F5F'--IVA TASA
			where c.ID = ALIAS_32.ID limit 1 ) POSIVA,
ALIAS_39.USUARIO USUARIO
	
FROM   V_TRFACTURAVENTA ALIAS_0

LEFT OUTER JOIN V_TREXTENSION ALIAS_2 ON ALIAS_0.TREXTENSION_ID = ALIAS_2.ID   
LEFT OUTER JOIN V_ESQUEMAOPERATIVO ALIAS_3 ON ALIAS_2.ESQUEMAOPERATIVO_ID = ALIAS_3.ID   
LEFT OUTER JOIN V_MOTIVOTRASLADO ALIAS_4 ON ALIAS_0.MOTIVOTRASLADO_ID = ALIAS_4.ID   

LEFT OUTER JOIN V_FLAG ALIAS_6 ON ALIAS_0.FLAG_ID = ALIAS_6.ID   
LEFT OUTER JOIN V_UNIDADOPERATIVA ALIAS_7 ON ALIAS_0.UNIDADOPERATIVA_ID = ALIAS_7.ID   
LEFT OUTER JOIN V_TIPODISTRIBUCIONCENCOS ALIAS_8 ON ALIAS_0.TIPODISTRIBUCIONCENCOS_ID = ALIAS_8.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_9 ON ALIAS_0.DISCRIMINADORCONTABLE4_ID = ALIAS_9.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_10 ON ALIAS_0.DISCRIMINADORCONTABLE5_ID = ALIAS_10.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_11 ON ALIAS_0.DISCRIMINADORCONTABLE6_ID = ALIAS_11.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_12 ON ALIAS_0.DISCRIMINADORCONTABLE7_ID = ALIAS_12.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_13 ON ALIAS_0.DISCRIMINADORCONTABLE8_ID = ALIAS_13.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_14 ON ALIAS_0.DISCRIMINADORCONTABLE9_ID = ALIAS_14.ID   
LEFT OUTER JOIN V_ITEMDISCCONTABLE ALIAS_15 ON ALIAS_0.DISCRIMINADORCONTABLE10_ID = ALIAS_15.ID   
LEFT OUTER JOIN V_PUNTOVENTA ALIAS_16 ON ALIAS_0.PUNTOVENTA_ID = ALIAS_16.ID   
LEFT OUTER JOIN V_TALONARIO ALIAS_17 ON ALIAS_0.TALONARIO_ID = ALIAS_17.ID   
LEFT OUTER JOIN V_AREARESPONSABILIDAD ALIAS_18 ON ALIAS_0.AREARESP_ID = ALIAS_18.ID   
LEFT OUTER JOIN V_DOMICILIO ALIAS_22 ON ALIAS_0.DOMICILIO_ID = ALIAS_22.ID   
LEFT OUTER JOIN V_PAIS ALIAS_19 ON ALIAS_22.PAIS_ID = ALIAS_19.ID   
LEFT OUTER JOIN V_PROVINCIA ALIAS_20 ON ALIAS_22.PROVINCIA_ID = ALIAS_20.ID   
LEFT OUTER JOIN V_CIUDAD ALIAS_21 ON ALIAS_22.CIUDAD_ID = ALIAS_21.ID  
LEFT OUTER JOIN V_TIPOTRANSACCION ALIAS_23 ON ALIAS_23.ID = ALIAS_0.TIPOTRANSACCION_ID  
LEFT OUTER JOIN V_ITEMFACTURAVENTA ALIAS_24 ON ALIAS_24.BO_PLACE_ID = ALIAS_0.ITEMSTRANSACCION_ID
LEFT OUTER JOIN V_MONEDA ALIAS_1 ON ALIAS_24.MONEDABASE_ID = ALIAS_1.ID   
LEFT OUTER JOIN V_MONEDA ALIAS_5 ON ALIAS_24.MONEDAORIGINAL_ID = ALIAS_5.ID     
LEFT OUTER JOIN V_IMPUTACIONCONTABLE ALIAS_25 ON ALIAS_25.ID = ALIAS_0.IMPUTACIONCONTABLE_ID  
LEFT OUTER JOIN V_CENTROCOSTOS ALIAS_26 ON ALIAS_26.ID = ALIAS_0.CENTROCOSTOS_ID  
LEFT OUTER JOIN V_PROYECTO ALIAS_27 ON ALIAS_27.ID = ALIAS_0.PROYECTO_ID  
LEFT OUTER JOIN V_UNIDADMEDIDA ALIAS_28 ON ALIAS_28.ID = ALIAS_24.UNIDADMEDIDA_ID  
LEFT OUTER JOIN V_TRANSACCION ALIAS_29 ON ALIAS_29.ID = ALIAS_0.VINCULOTR_ID  
LEFT OUTER JOIN V_UNIDADNEGOCIO ALIAS_30 ON ALIAS_30.ID = ALIAS_2.UNIDADNEGOCIO_ID  
LEFT OUTER JOIN V_TIPOPAGO ALIAS_31 ON ALIAS_31.ID = ALIAS_0.TIPOPAGO_ID  
LEFT OUTER JOIN V_OPERADORCOMERCIAL ALIAS_32 ON ALIAS_32.ID = ALIAS_0.destinatario_ID  
LEFT OUTER JOIN V_SEGMENTO ALIAS_33 ON ALIAS_33.ID = ALIAS_32.SEGMENTO_ID  
LEFT OUTER JOIN V_CONCEPTOCOMERCIAL ALIAS_34 ON ALIAS_34.ID = ALIAS_24.REFERENCIA_ID  
LEFT OUTER JOIN V_SEGMENTO ALIAS_35 ON ALIAS_35.ID = ALIAS_34.SEGMENTO_ID  
LEFT OUTER JOIN UD_TRFACTURAVENTA ALIAS_36 ON ALIAS_36.ID = ALIAS_0.BOEXTENSION_ID  
LEFT OUTER JOIN V_ATRIBUTOSAUXILIARES ALIAS_37 ON ALIAS_0.ATRIBUTOSAUXILIARES_ID = ALIAS_37.ID
LEFT OUTER JOIN V_TRORDENVENTA ALIAS_38 ON ALIAS_38.NOMBRE = ALIAS_37.NOMBREATRIBUTOAUXILIAR1
LEFT OUTER JOIN V_AUDITORIA ALIAS_39 ON ALIAS_38.ID = ALIAS_39.OBJETO_ID



WHERE ALIAS_0.BO_PLACE_ID = '{89C23511-3F01-11D5-86AD-0080AD403F5F}'   
AND   0=0  
AND  ALIAS_0.TIPOTRANSACCION_ID = '{89C235A5-3F01-11D5-86AD-0080AD403F5F}'  
AND ALIAS_0.ESTADO LIKE 'C'
AND ALIAS_39.ACCION LIKE 'Alta'

ORDER BY ALIAS_0.FECHAACTUAL DESC, ALIAS_0.NUMERODOCUMENTO DESC
"""

cabeceras = ["Unidad Operativa", "Codigo Esq.op.", "Esquema Op.", "Num. Factura", "Cuit cliente", "Fecha facturado",
             "Estado", "Cod. estado", "Vendedor", "Cliente", "Producto", "Unidad de negocio", "Tipo de pago", "Tipo de cliente",
             "Estado de cliente", "Zona", "Familia", "Estado de producto", "Cantidad de producto", "Valor de producto", "Subtotal",
             "Precio Unitario Final", "Porcentaje de Bonificiación", "Valor total", "Numero CUIT", "Codigo de cliente", "Posición de IVA", "Usuario"]


#Modificar el filepath en caso de ser necesario, siempre con doble barra \\, para evitar errores de sintaxis
guardar_en = "path"


#Revisar el script pgconnection en caso de necesitar modificar datos de conexion
conn = pgconnection.get_connection("NOMBREBD")

export_to_excel(conn, query, cabeceras, guardar_en)
